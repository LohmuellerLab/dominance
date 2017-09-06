#!/usr/bin/env python2
# -*- coding: utf-8 -*-
"""
Created on Tue Sep 5 2017

@author: Christian D. Huber
"""

# import sys
# sys.path.append("optional path to Dominance_onePop, Dominance_twoPop, and dadi_wrapper_functions")
import Dominance_onePop, Dominance_twoPop
import dadi_R_wrapper_functions
import numpy
import dadi
import pickle
import pylab


#############################################################################
print "Load SFS data"

f1 = open("Lyrata_SFS.txt", "r")
raw1 = f1.readlines()
f1.close()
sampleSize1 = int(raw1[1])

sfs_NS1 = numpy.array(raw1[5][:-1].split(" "), dtype='float64').tolist()
lensfs = len(sfs_NS1)
sfs_NS1 = [0.] + sfs_NS1 + numpy.zeros(shape=(sampleSize1-lensfs,)).tolist()
sfs_NS1_noMask = dadi.Spectrum(sfs_NS1, data_folded=True, mask = [True]*1 + [False]*(lensfs) + [True]*(sampleSize1-lensfs))  

sfs_S1 = numpy.array(raw1[3][:-1].split(" "), dtype='float64').tolist()
lensfs = len(sfs_S1)
sfs_S1 = [0.] + sfs_S1 + numpy.zeros(shape=(sampleSize1-lensfs,)).tolist()
sfs_S1_noMask = dadi.Spectrum(sfs_S1, data_folded=True, mask = [True]*1 + [False]*(lensfs) + [True]*(sampleSize1-lensfs))  

f2 = open("Thaliana_SFS.txt", "r")
raw2 = f2.readlines()
f2.close()
sampleSize2 = int(raw2[1])

sfs_NS2 = numpy.array(raw2[5][:-1].split(" "), dtype='float64').tolist()
lensfs = len(sfs_NS2)
sfs_NS2 = [0.] + sfs_NS2 + numpy.zeros(shape=(sampleSize2-lensfs,)).tolist()
sfs_NS2_noMask = dadi.Spectrum(sfs_NS2, data_folded=True, mask = [True]*1 + [False]*(lensfs) + [True]*(sampleSize2-lensfs))  # Mask singletons!!

sfs_S2 = numpy.array(raw2[3][:-1].split(" "), dtype='float64').tolist()
lensfs = len(sfs_S2)
sfs_S2 = [0.] + sfs_S2 + numpy.zeros(shape=(sampleSize2-lensfs,)).tolist()
sfs_S2_noMask = dadi.Spectrum(sfs_S2, data_folded=True, mask = [True]*1 + [False]*(lensfs) + [True]*(sampleSize2-lensfs))  # Mask singletons!!


#############################################################################
print "Estimate demography -- Three epoch model"

if True:
    
    lower_bound=[0.0001, 0, 0.0001, 0]
    upper_bound=[100, 2, 10, 2]

    nrep = 1000
    
    popt_1_array = sorted([dadi.Inference.optimize_log(numpy.random.uniform(lower_bound, upper_bound), sfs_S1_noMask, dadi_R_wrapper_functions.three_epoch_ex, [100,120,140], lower_bound=lower_bound, upper_bound=upper_bound, verbose=1, maxiter=10, full_output=True)[:2][::-1] for i in range(nrep)], key=lambda pair: pair[0], reverse = False)  
    popt_1 = popt_1_array[0][1]
    pickle.dump([popt_1_array, popt_1], open('demo_lyrata','wb'))
   
    lower_bound=[0.001, 0, 0.001, 0]
    upper_bound=[10, 2, 10, 2]

    popt_2_array = sorted([dadi.Inference.optimize_log(numpy.random.uniform(lower_bound, upper_bound), sfs_S2_noMask, dadi_R_wrapper_functions.three_epoch_ex, [100,120,140], lower_bound=lower_bound, upper_bound=upper_bound, verbose=1, maxiter=10, full_output=True)[:2][::-1] for i in range(nrep)], key=lambda pair: pair[0], reverse = False)                         
    popt_2 = popt_2_array[0][1]
    pickle.dump([popt_2_array, popt_2], open('demo_thaliana','wb'))
    
popt_1_array, popt_1 = pickle.load(open('demo_lyrata','rb'))
popt_2_array, popt_2 = pickle.load(open('demo_thaliana','rb'))

theta_S1 = dadi.Inference.optimal_sfs_scaling(dadi_R_wrapper_functions.three_epoch_ex(popt_1, sfs_S1_noMask.sample_sizes, [100,120,140]), sfs_S1_noMask)
theta_NS1 = 2.31*theta_S1   # ratio 2.31 for nonsynonymous to synonymous mutation rate
L1 = float(raw1[7])/(2.31+1)   # Effective sequence length for synonymous mutations
NeAnc1 = theta_S1/(4*L1*0.7e-8)

theta_S2 = dadi.Inference.optimal_sfs_scaling(dadi_R_wrapper_functions.three_epoch_ex(popt_2, sfs_S2_noMask.sample_sizes, [100,120,140]), sfs_S2_noMask)
theta_NS2 = 2.31*theta_S2   # ratio 2.31 for nonsynonymous to synonymous mutation rate
L2 = float(raw2[7])/(2.31+1)   # Effective sequence length for synonymous mutations
NeAnc2 = theta_S2/(4*L2*0.7e-8)

pickle.dump([sfs_NS1_noMask, sfs_S1_noMask, sfs_NS2_noMask, sfs_S2_noMask, L1, L2, NeAnc1, NeAnc2], open('sfs_data','wb'))

if True:
    # Comparing expected vs. observed synonymous SFS:
    
    pylab.rcParams.update({'font.size': 22})
    pylab.figure(figsize=(12, 5))
    
    pylab.subplot(1,2,1)
    pylab.plot(sfs_S1_noMask, '--ob', mfc='grey', zorder=-100, label="Synonymous SFS")
    pylab.plot(dadi_R_wrapper_functions.three_epoch_ex(popt_1, sfs_S1_noMask.sample_sizes, [100,120,140]).fold()*theta_S1, '--ob', mfc='red', zorder=-1000, label="Three-epoch model")
    pylab.xlim([0,12])
    pylab.ylim([0,350000])
    pylab.xlabel('MAF')
    pylab.ylabel('Number of SNPs')
    pylab.legend(prop={'size': 16}, frameon=False, title="A. lyrata")
        
    pylab.subplot(1,2,2)
    pylab.plot(sfs_S2_noMask, '--ob', mfc='grey', zorder=-100, label="Synonymous SFS")
    pylab.plot(dadi_R_wrapper_functions.three_epoch_ex(popt_2, sfs_S2_noMask.sample_sizes, [100,120,140]).fold()*theta_S2, '--ob', mfc='red', zorder=-1000, label="Three-epoch model")
    pylab.xlim([0,9])
    pylab.ylim([0,35000])
    pylab.xlabel('MAF')
    pylab.ylabel('Number of SNPs')
    pylab.legend(prop={'size': 16}, frameon=False, title="A. thaliana")
    
    pylab.tight_layout()

    
#############################################################################
print "Computing cache:"

if True:    # Note that this can take several hours
    # Set last parameter (num_cores) to change the number of computing cores
    Dominance_onePop.generatecache(tuple(popt_1.tolist()), sfs_S1_noMask.sample_sizes, dadi_R_wrapper_functions.three_epoch_selection, note="s1", pts=1000, Npts_large = 1000, Npts_h_large = 1001, NeAnc=NeAnc1, num_cores = 6)
    Dominance_onePop.generatecache(tuple(popt_2.tolist()), sfs_S2_noMask.sample_sizes, dadi_R_wrapper_functions.three_epoch_selection, note="s2", pts=1000, Npts_large = 1000, Npts_h_large = 1001, NeAnc=NeAnc2, num_cores = 6)                      

spectra_cached = dict(s1 = Dominance_onePop.loadSpectraObj('s1'), s2 = Dominance_onePop.loadSpectraObj('s2'))    


#############################################################################
print "Estimating DFE and h params -- outcrossing species:"    

lower_bound=[0, 10, 0, 0]
upper_bound=[1, 5000, 1, 1]

def demo_selection_dist(params, ns, sel_dist, theta):
    return Dominance_onePop.demo_selection_dist(params, ns, sel_dist, theta, cache=spectra_cached["s1"])
        
if True:
    result_1_additive = [Dominance_onePop.optimize_log(numpy.random.uniform(lower_bound, upper_bound), sfs_NS1_noMask, demo_selection_dist, Dominance_onePop.gamma_dist, theta_NS1, lower_bound=lower_bound, upper_bound=upper_bound, maxiter=100, verbose=1, fixed_params=[None, None, 0.5, 0]) for i in range(1000)]
    result_1_dominance = [Dominance_onePop.optimize_log(numpy.random.uniform(lower_bound, upper_bound), sfs_NS1_noMask, demo_selection_dist, Dominance_onePop.gamma_dist, theta_NS1, lower_bound=lower_bound, upper_bound=upper_bound, maxiter=100, verbose=1, fixed_params=[None, None, None, 0]) for i in range(1000)]
    result_1_dominanceFunc = [Dominance_onePop.optimize_log(numpy.random.uniform(lower_bound, upper_bound), sfs_NS1_noMask, demo_selection_dist, Dominance_onePop.gamma_dist, theta_NS1, lower_bound=lower_bound, upper_bound=upper_bound, maxiter=100, verbose=1, fixed_params=[None, None, None, None]) for i in range(1000)]
                             
    # Sort accord to LL:
    
    result_1_additive = sorted(result_1_additive, key=lambda pair: pair[0], reverse = True)                                                 
    result_1_dominance = sorted(result_1_dominance, key=lambda pair: pair[0], reverse = True)                                                 
    result_1_dominanceFunc = sorted(result_1_dominanceFunc, key=lambda pair: pair[0], reverse = True)                                                      
                 
    pickle.dump([result_1_additive, result_1_dominance, result_1_dominanceFunc], open('res_1','wb'))
result_1_additive, result_1_dominance, result_1_dominanceFunc = pickle.load(open('res_1','rb'))


#############################################################################                          
print "Estimating DFE and h params -- selfing species:"

lower_bound=[0, 10, 0, 0]
upper_bound=[1, 5000, 1, 1]

def demo_selection_dist(params, ns, sel_dist, theta):
    return Dominance_onePop.demo_selection_dist(params, ns, sel_dist, theta, cache=spectra_cached["s2"])
  
if True:
    result_2_additive = [Dominance_onePop.optimize_log(numpy.random.uniform(lower_bound, upper_bound), sfs_NS2_noMask, demo_selection_dist, Dominance_onePop.gamma_dist, theta_NS2, lower_bound=lower_bound, upper_bound=upper_bound, maxiter=100, verbose=1, fixed_params=[None, None, 0.5, 0]) for i in range(1000)]
    result_2_dominance = [Dominance_onePop.optimize_log(numpy.random.uniform(lower_bound, upper_bound), sfs_NS2_noMask, demo_selection_dist, Dominance_onePop.gamma_dist, theta_NS2, lower_bound=lower_bound, upper_bound=upper_bound, maxiter=100, verbose=1, fixed_params=[None, None, None, 0]) for i in range(1000)]
    result_2_dominanceFunc = [Dominance_onePop.optimize_log(numpy.random.uniform(lower_bound, upper_bound), sfs_NS2_noMask, demo_selection_dist, Dominance_onePop.gamma_dist, theta_NS2, lower_bound=lower_bound, upper_bound=upper_bound, maxiter=100, verbose=1, fixed_params=[None, None, None, None]) for i in range(1000)]
    
    # Sort accord to LL: 
        
    result_2_additive = sorted(result_2_additive, key=lambda pair: pair[0], reverse = True)                                                 
    result_2_dominance = sorted(result_2_dominance, key=lambda pair: pair[0], reverse = True)                                                 
    result_2_dominanceFunc = sorted(result_2_dominanceFunc, key=lambda pair: pair[0], reverse = True)                                                 
    
    pickle.dump([result_2_additive, result_2_dominance, result_2_dominanceFunc], open('res_2','wb'))
result_2_additive, result_2_dominance, result_2_dominanceFunc = pickle.load(open('res_2','rb'))


#############################################################################
print("Estimating DFE and h params -- use both outrossing and selfing species:")

lower_bound=[0, 0.00001, 0, 0]
upper_bound=[1, 1, 1, 1]

if True:
    result_comb_additive = sorted([Dominance_twoPop.optimize_log(numpy.random.uniform(lower_bound, upper_bound), sfs_NS1_noMask, sfs_NS2_noMask, spectra_cached["s1"], spectra_cached["s2"], Dominance_onePop.demo_selection_dist, Dominance_onePop.gamma_dist, NeAnc1, NeAnc2*2, theta_NS1, theta_NS2, lower_bound=lower_bound, upper_bound=upper_bound, fixed_params1=[None, None, 0.5, 0.0], fixed_params2=[None, None, 0.5, 0.0], verbose=1, maxiter=100) for i in range(1000)], key=lambda pair: pair[0], reverse = True)
    result_comb_const_h = sorted([Dominance_twoPop.optimize_log(numpy.random.uniform(lower_bound, upper_bound), sfs_NS1_noMask, sfs_NS2_noMask, spectra_cached["s1"], spectra_cached["s2"], Dominance_onePop.demo_selection_dist, Dominance_onePop.gamma_dist, NeAnc1, NeAnc2*2, theta_NS1, theta_NS2, lower_bound=lower_bound, upper_bound=upper_bound, fixed_params1=[None, None, None, 0.0], fixed_params2=[None, None, 0.5, 0.0], verbose=1, maxiter=100) for i in range(1000)], key=lambda pair: pair[0], reverse = True)
    result_comb_dominanceFunc = sorted([Dominance_twoPop.optimize_log(numpy.random.uniform(lower_bound, upper_bound), sfs_NS1_noMask, sfs_NS2_noMask, spectra_cached["s1"], spectra_cached["s2"], Dominance_onePop.demo_selection_dist, Dominance_onePop.gamma_dist, NeAnc1, NeAnc2*2, theta_NS1, theta_NS2, lower_bound=lower_bound, upper_bound=upper_bound, fixed_params1=[None, None, None, None], fixed_params2=[None, None, 0.5, 0.0], verbose=1, maxiter=100) for i in range(1000)], key=lambda pair: pair[0], reverse = True)
    pickle.dump([result_comb_additive, result_comb_const_h, result_comb_dominanceFunc], open('res_combined','wb'))

   
result_comb_additive, result_comb_const_h, result_comb_dominanceFunc = pickle.load(open('res_combined','rb'))


#############################################################################
print("Plots: compare nonsynonymous SFS from data with expected SFS from dominance models")

pylab.figure(figsize=(18, 12))
pylab.subplot(2,3,1)
pylab.plot(Dominance_onePop.demo_selection_dist(tuple(result_comb_additive[0][1]*[1,1*NeAnc1,1,1]), sfs_S1.sample_sizes, Dominance_onePop.gamma_dist, theta_NS1, cache=spectra_cached["s1"]).fold(), '--ob', mfc='blue', zorder=-1000)
pylab.plot(sfs_NS1_noMask, '--ob', mfc='grey', zorder=-100)
pylab.subplot(2,3,2)
pylab.plot(Dominance_onePop.demo_selection_dist(tuple(result_comb_const_h[0][1]*[1,1*NeAnc1,1,1]), sfs_S1.sample_sizes, Dominance_onePop.gamma_dist, theta_NS1, cache=spectra_cached["s1"]).fold(), '--ob', mfc='red', zorder=-1000)
pylab.plot(sfs_NS1_noMask, '--ob', mfc='grey', zorder=-100)
pylab.subplot(2,3,3)
pylab.plot(Dominance_onePop.demo_selection_dist(tuple(result_comb_dominanceFunc[0][1]*[1,1*NeAnc1,1,1]), sfs_S1.sample_sizes, Dominance_onePop.gamma_dist, theta_NS1, cache=spectra_cached["s1"]).fold(), '--ob', mfc='violet', zorder=-1000)
pylab.plot(sfs_NS1_noMask, '--ob', mfc='grey', zorder=-100)

pylab.subplot(2,3,4)
pylab.plot(Dominance_onePop.demo_selection_dist(tuple(numpy.append(result_comb_additive[0][1][:2]*[1,2*NeAnc2], [0.5,0.])), sfs_S2.sample_sizes, Dominance_onePop.gamma_dist, theta_NS2, cache=spectra_cached["s2"]).fold(), '--ob', mfc='blue', zorder=-1000)
pylab.plot(sfs_NS2_noMask, '--ob', mfc='grey', zorder=-100)
pylab.subplot(2,3,5)
pylab.plot(Dominance_onePop.demo_selection_dist(tuple(numpy.append(result_comb_const_h[0][1][:2]*[1,2*NeAnc2], [0.5,0.])), sfs_S2.sample_sizes, Dominance_onePop.gamma_dist, theta_NS2, cache=spectra_cached["s2"]).fold(), '--ob', mfc='red', zorder=-1000)
pylab.plot(sfs_NS2_noMask, '--ob', mfc='grey', zorder=-100)
pylab.subplot(2,3,6)
pylab.plot(Dominance_onePop.demo_selection_dist(tuple(numpy.append(result_comb_dominanceFunc[0][1][:2]*[1,2*NeAnc2], [0.5,0.])), sfs_S2.sample_sizes, Dominance_onePop.gamma_dist, theta_NS2, cache=spectra_cached["s2"]).fold(), '--ob', mfc='violet', zorder=-1000)
pylab.plot(sfs_NS2_noMask, '--ob', mfc='grey', zorder=-100)
pylab.show()
 

#############################################################################
print("Write results into excel file named results_table_dominance.xlsx")

import openpyxl

wb = openpyxl.Workbook()
ws = wb.active

ws['A4'].value = "Outcrossing"

ws['A5'].value = "Model"
ws['B5'].value = "Alpha"
ws['C5'].value = "Beta"
ws['D5'].value = "h_intercept"
ws['E5'].value = "h_rate"
ws['F5'].value = "LL"
ws['A6'].value = "Additive"
ws['A7'].value = "Constant h"
ws['A8'].value = "h-s Relationship"

ws['A6'].value = "Additive"
ws['B6'].value = result_1_additive[0][1][0]
ws['C6'].value = result_1_additive[0][1][1]/NeAnc1
ws['D6'].value = result_1_additive[0][1][2]
ws['E6'].value = result_1_additive[0][1][3]*NeAnc1
ws['F6'].value = result_1_additive[0][0]
ws['A7'].value = "Constant h"
ws['B7'].value = result_1_dominance[0][1][0]
ws['C7'].value = result_1_dominance[0][1][1]/NeAnc1
ws['D7'].value = result_1_dominance[0][1][2]
ws['E7'].value = result_1_dominance[0][1][3]*NeAnc1
ws['F7'].value = result_1_dominance[0][0]
ws['A8'].value = "h-s Relationship"
ws['B8'].value = result_1_dominanceFunc[0][1][0]
ws['C8'].value = result_1_dominanceFunc[0][1][1]/NeAnc1
ws['D8'].value = result_1_dominanceFunc[0][1][2]
ws['E8'].value = result_1_dominanceFunc[0][1][3]*NeAnc1
ws['F8'].value = result_1_dominanceFunc[0][0]

ws['A10'].value = "Inbreeding"

ws['A11'].value = "Model"
ws['B11'].value = "Alpha"
ws['C11'].value = "Beta"
ws['D11'].value = "h_intercept"
ws['E11'].value = "h_rate"
ws['F11'].value = "LL"
ws['A12'].value = "Additive"
ws['A13'].value = "Constant h"
ws['A14'].value = "h-s Relationship"

ws['A12'].value = "Additive"
ws['B12'].value = result_2_additive[0][1][0]
ws['C12'].value = result_2_additive[0][1][1]/NeAnc2/2  # factor of 2 acount for fact that inference assumes outcrossing
ws['D12'].value = result_2_additive[0][1][2]
ws['E12'].value = result_2_additive[0][1][3]*NeAnc2
ws['F12'].value = result_2_additive[0][0]
ws['A13'].value = "Constant h"
ws['B13'].value = result_2_dominance[0][1][0]
ws['C13'].value = result_2_dominance[0][1][1]/NeAnc2/2
ws['D13'].value = result_2_dominance[0][1][2]
ws['E13'].value = result_2_dominance[0][1][3]*NeAnc2
ws['F13'].value = result_2_dominance[0][0]
ws['A14'].value = "h-s Relationship"
ws['B14'].value = result_2_dominanceFunc[0][1][0]
ws['C14'].value = result_2_dominanceFunc[0][1][1]/NeAnc2/2
ws['D14'].value = result_2_dominanceFunc[0][1][2]
ws['E14'].value = result_2_dominanceFunc[0][1][3]*NeAnc2
ws['F14'].value = result_2_dominanceFunc[0][0]

ws['A16'].value = "Combined"

ws['A17'].value = "Model"
ws['B17'].value = "Alpha"
ws['C17'].value = "Beta"
ws['D17'].value = "h_intercept"
ws['E17'].value = "h_rate"
ws['F17'].value = "LL"
ws['A18'].value = "Additive"
ws['A19'].value = "Constant h"
ws['A20'].value = "h-s Relationship"

ws['A18'].value = "Additive"
ws['B18'].value = result_comb_additive[0][1][0]
ws['C18'].value = result_comb_additive[0][1][1]
ws['D18'].value = result_comb_additive[0][1][2]
ws['E18'].value = result_comb_additive[0][1][3]*NeAnc1
ws['F18'].value = result_comb_additive[0][0]
ws['A19'].value = "Constant h"
ws['B19'].value = result_comb_const_h[0][1][0]
ws['C19'].value = result_comb_const_h[0][1][1]
ws['D19'].value = result_comb_const_h[0][1][2]
ws['E19'].value = result_comb_const_h[0][1][3]*NeAnc1
ws['F19'].value = result_comb_const_h[0][0]
ws['A20'].value = "h-s Relationship"
ws['B20'].value = result_comb_dominanceFunc[0][1][0]
ws['C20'].value = result_comb_dominanceFunc[0][1][1]
ws['D20'].value = result_comb_dominanceFunc[0][1][2]
ws['E20'].value = result_comb_dominanceFunc[0][1][3]*NeAnc1
ws['F20'].value = result_comb_dominanceFunc[0][0]

wb.save("results_table_dominance.xlsx")

  